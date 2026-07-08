from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from spectrail.core.models import RequirementIR


HEADERS = [
    "ID",
    "Title",
    "Type",
    "EARS Pattern",
    "Statement",
    "Subject",
    "Condition",
    "Response",
    "Priority",
    "Verification Method",
    "Confidence",
    "Review Status",
    "Source Section",
    "Source Block ID",
    "Source Quote",
    "Source Match Status",
    "Tags",
]


def export_requirements_xlsx(requirements: list[RequirementIR], path: str | Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Requirements"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="E8EEF7")

    for requirement in requirements:
        source = requirement.sources[0] if requirement.sources else None
        sheet.append(
            [
                requirement.id,
                requirement.title,
                requirement.type,
                requirement.ears_pattern,
                requirement.statement,
                requirement.subject,
                requirement.condition,
                requirement.response,
                requirement.priority,
                requirement.verification_method,
                requirement.confidence,
                requirement.review_status,
                source.section if source else None,
                source.block_id if source else None,
                source.quote if source else None,
                source.match_status if source else None,
                ", ".join(requirement.tags),
            ]
        )

    widths = [12, 20, 18, 18, 42, 16, 28, 32, 12, 22, 12, 16, 28, 16, 46, 20, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
