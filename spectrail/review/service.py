from __future__ import annotations

import os
import shutil
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import TypeAdapter

from spectrail.core.io import (
    read_json,
    read_reqir_package,
    read_reqir_items,
    reqir_package_dump,
    write_json,
)
from spectrail.core.models import ReqIRPackage, RequirementIR
from spectrail.exporters.xlsx_exporter import export_requirements_xlsx
from spectrail.review.review_log import collect_review_log
from spectrail.review.review_state import apply_review_action
from spectrail.task_transactions import task_operation, task_root_for_artifact


ReqListAdapter = TypeAdapter(list[RequirementIR])


def load_requirements(path: str | Path) -> list[RequirementIR]:
    return ReqListAdapter.validate_python(read_reqir_items(path))


def load_requirement_package(path: str | Path) -> ReqIRPackage:
    return read_reqir_package(path)


def refresh_review_package(
    reqir_path: str | Path,
    review_log_path: str | Path,
    xlsx_path: str | Path,
    requirements: list[RequirementIR],
    *,
    metadata: dict[str, Any] | None = None,
) -> None:
    task_root = task_root_for_artifact(reqir_path)
    if task_root is None:
        return _refresh_review_package_locked(
            reqir_path,
            review_log_path,
            xlsx_path,
            requirements,
            metadata=metadata,
        )
    with task_operation(task_root, "review_refresh"):
        return _refresh_review_package_locked(
            reqir_path,
            review_log_path,
            xlsx_path,
            requirements,
            metadata=metadata,
        )


def _refresh_review_package_locked(
    reqir_path: str | Path,
    review_log_path: str | Path,
    xlsx_path: str | Path,
    requirements: list[RequirementIR],
    *,
    metadata: dict[str, Any] | None = None,
) -> None:
    reqir_target = Path(reqir_path)
    review_log_target = Path(review_log_path)
    xlsx_target = Path(xlsx_path)
    task_root = task_root_for_artifact(reqir_target)
    staging_parent = task_root or reqir_target.parent
    staging_parent.mkdir(parents=True, exist_ok=True)
    review_log_payload = collect_review_log(requirements)
    reqir_payload = reqir_package_dump(
        requirements,
        metadata={
            **(metadata or {}),
            "export_state": "review_snapshot",
        },
    )
    with tempfile.TemporaryDirectory(
        prefix=".review_prepare_",
        dir=staging_parent,
    ) as temporary:
        staging_root = Path(temporary)
        staged_review_log = staging_root / "review_log.json"
        staged_reqir = staging_root / "reqir.json"
        staged_xlsx = staging_root / "requirements.xlsx"
        write_json(staged_review_log, review_log_payload)
        write_json(staged_reqir, reqir_payload)
        export_requirements_xlsx(requirements, staged_xlsx)
        _validate_staged_review_artifacts(
            staged_review_log=staged_review_log,
            staged_reqir=staged_reqir,
            staged_xlsx=staged_xlsx,
            expected_review_log=review_log_payload,
            expected_requirement_ids=[
                requirement.id for requirement in requirements
            ],
        )
        _publish_review_artifacts(
            staging_root=staging_root,
            artifacts=[
                (staged_review_log, review_log_target),
                (staged_reqir, reqir_target),
                (staged_xlsx, xlsx_target),
            ],
        )


def _validate_staged_review_artifacts(
    *,
    staged_review_log: Path,
    staged_reqir: Path,
    staged_xlsx: Path,
    expected_review_log: list[dict[str, Any]],
    expected_requirement_ids: list[str],
) -> None:
    if read_json(staged_review_log) != expected_review_log:
        raise ValueError("staged review log does not match review state")
    package = read_reqir_package(staged_reqir)
    if [item.id for item in package.items] != expected_requirement_ids:
        raise ValueError("staged ReqIR package does not match review state")
    workbook = load_workbook(staged_xlsx, read_only=True, data_only=True)
    try:
        if "Requirements" not in workbook.sheetnames:
            raise ValueError("staged review workbook is missing Requirements")
        if workbook["Requirements"].max_row != len(expected_requirement_ids) + 1:
            raise ValueError("staged review workbook row count is invalid")
    finally:
        workbook.close()


def _publish_review_artifacts(
    *,
    staging_root: Path,
    artifacts: list[tuple[Path, Path]],
) -> None:
    backup_root = staging_root / "backup"
    backup_root.mkdir()
    backups: list[tuple[Path, Path, bool]] = []
    for index, (_, target) in enumerate(artifacts):
        target.parent.mkdir(parents=True, exist_ok=True)
        existed = target.exists()
        backup = backup_root / f"{index:02d}-{target.name}"
        if existed:
            shutil.copy2(target, backup)
            _fsync_file(backup)
        backups.append((target, backup, existed))
    _fsync_directory(backup_root)
    for staged, _ in artifacts:
        _fsync_file(staged)
    _fsync_directory(staging_root)

    try:
        for staged, target in artifacts:
            os.replace(staged, target)
            _fsync_directory(target.parent)
    except BaseException as publication_error:
        rollback_errors: list[OSError] = []
        for target, backup, existed in backups:
            try:
                if existed:
                    os.replace(backup, target)
                    _fsync_directory(target.parent)
                elif target.exists():
                    target.unlink()
                    _fsync_directory(target.parent)
            except OSError as rollback_error:
                rollback_errors.append(rollback_error)
        if rollback_errors:
            raise RuntimeError(
                "review publication failed and rollback was incomplete"
            ) from publication_error
        raise


def _fsync_file(path: Path) -> None:
    with path.open("rb") as handle:
        os.fsync(handle.fileno())


def _fsync_directory(path: Path) -> None:
    try:
        descriptor = os.open(path, os.O_RDONLY)
    except OSError:
        return
    try:
        os.fsync(descriptor)
    except OSError:
        pass
    finally:
        os.close(descriptor)


def apply_review_to_package(
    reqir_path: str | Path,
    review_log_path: str | Path,
    xlsx_path: str | Path,
    requirement_id: str,
    action: str,
    patch: dict[str, Any] | None = None,
    reviewer: str | None = None,
    reason: str | None = None,
) -> RequirementIR:
    task_root = task_root_for_artifact(reqir_path)
    if task_root is None:
        return _apply_review_to_package_locked(
            reqir_path=reqir_path,
            review_log_path=review_log_path,
            xlsx_path=xlsx_path,
            requirement_id=requirement_id,
            action=action,
            patch=patch,
            reviewer=reviewer,
            reason=reason,
        )
    with task_operation(task_root, "review_apply"):
        return _apply_review_to_package_locked(
            reqir_path=reqir_path,
            review_log_path=review_log_path,
            xlsx_path=xlsx_path,
            requirement_id=requirement_id,
            action=action,
            patch=patch,
            reviewer=reviewer,
            reason=reason,
        )


def _apply_review_to_package_locked(
    reqir_path: str | Path,
    review_log_path: str | Path,
    xlsx_path: str | Path,
    requirement_id: str,
    action: str,
    patch: dict[str, Any] | None = None,
    reviewer: str | None = None,
    reason: str | None = None,
) -> RequirementIR:
    package = load_requirement_package(reqir_path)
    requirements = package.items
    target = next((req for req in requirements if req.id == requirement_id), None)
    if target is None:
        raise ValueError(f"requirement not found: {requirement_id}")

    apply_review_action(
        target,
        action,
        patch=patch,
        reviewer=reviewer,
        reason=reason,
    )
    refresh_review_package(
        reqir_path,
        review_log_path,
        xlsx_path,
        requirements,
        metadata=package.metadata,
    )
    return target
