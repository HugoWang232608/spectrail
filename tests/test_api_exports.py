from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def test_api_exports_download_files(api_client: TestClient, completed_api_task: dict):
    task_id = completed_api_task["task_id"]

    reqir = api_client.get(
        f"/api/tasks/{task_id}/exports/reqir.json"
        "?expected_run_generation=1"
    )
    assert reqir.status_code == 200
    assert reqir.headers["content-type"].startswith("application/json")
    assert reqir.headers["cache-control"] == "private, no-store"
    assert reqir.headers["x-spectrail-run-generation"] == "1"
    assert len(reqir.json()["items"]) >= 14

    xlsx = api_client.get(
        f"/api/tasks/{task_id}/exports/requirements.xlsx"
        "?expected_run_generation=1"
    )
    assert xlsx.status_code == 200
    assert xlsx.headers["x-spectrail-run-generation"] == "1"
    workbook = load_workbook(BytesIO(xlsx.content))
    assert workbook["Requirements"].max_row >= 15


def test_api_export_before_completion_returns_409(api_client: TestClient):
    created = api_client.post("/api/tasks", json={})
    task_id = created.json()["task_id"]

    response = api_client.get(
        f"/api/tasks/{task_id}/exports/requirements.xlsx"
        "?expected_run_generation=0"
    )
    assert response.status_code == 409
    assert response.json()["detail"]["code"] == "TASK_NOT_COMPLETED"


def test_api_exports_reject_stale_task_generation(
    api_client: TestClient,
    completed_api_task: dict,
):
    task_id = completed_api_task["task_id"]
    store = api_client.app.state.task_store
    store.begin_run(task_id)
    store.update_task(task_id, status="completed")

    for filename in ("reqir.json", "requirements.xlsx"):
        response = api_client.get(
            f"/api/tasks/{task_id}/exports/{filename}"
            "?expected_run_generation=1"
        )
        assert response.status_code == 409
        assert response.json()["detail"]["code"] == "RUN_GENERATION_CHANGED"
