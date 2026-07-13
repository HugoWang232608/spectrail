from pathlib import Path

from openpyxl import load_workbook

from spectrail.cli import main
from spectrail.core.io import read_json


def test_mock_pipeline_generates_p0_outputs(tmp_path: Path):
    output = tmp_path / "demo"
    assert main(["extract", "docs/sample_srs.md", "--model-mode", "mock", "--output", str(output)]) == 0

    expected = [
        output / "plan.json",
        output / "run_manifest.json",
        output / "parsed" / "blocks.json",
        output / "parsed" / "evidence_index.json",
        output / "extracted" / "reqir.raw.json",
        output / "extracted" / "quote_matches.json",
        output / "extracted" / "reqir.validated.json",
        output / "extracted" / "source_map.json",
        output / "extracted" / "validation_report.json",
        output / "extracted" / "source_locator_report.json",
        output / "extracted" / "source_locator_failures.json",
        output / "review" / "review_log.json",
        output / "exports" / "reqir.json",
        output / "exports" / "requirements.xlsx",
    ]
    for path in expected:
        assert path.exists(), path

    package = read_json(output / "extracted" / "reqir.validated.json")
    requirements = package["items"]
    assert len(requirements) >= 14

    plan = read_json(output / "plan.json")
    assert plan["planner"] == "fixed_workflow_v1"

    run_manifest = read_json(output / "run_manifest.json")
    assert run_manifest["status"] == "completed"
    assert run_manifest["counts"]["validated_requirements"] >= 14
    assert run_manifest["evidence"]["block_count"] > 0
    assert run_manifest["evidence"]["quote_match_count"] > 0
    assert run_manifest["evidence"]["policy"] == "structured_if_available"

    for requirement in requirements:
        assert requirement["id"]
        assert requirement["statement"]
        assert requirement["review_status"] == "pending"
        assert requirement["sources"]
        assert all(source["source_evidence_key"] for source in requirement["sources"])
        assert all(source["text_locator"] for source in requirement["sources"])
        assert any(
            source["match_status"] in {"PASS_EXACT", "PASS_NORMALIZED"}
            for source in requirement["sources"]
        )

    workbook = load_workbook(output / "exports" / "requirements.xlsx")
    assert workbook["Requirements"].max_row == len(requirements) + 1
