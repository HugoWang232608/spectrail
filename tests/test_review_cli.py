from pathlib import Path

from openpyxl import load_workbook

from spectrail.cli import main
from spectrail.core.io import read_json, write_json


def test_review_cli_refresh_without_action_preserves_package_metadata(tmp_path: Path):
    output = tmp_path / "demo"
    assert (
        main(
            [
                "extract",
                "docs/sample_srs.md",
                "--model-mode",
                "mock",
                "--output",
                str(output),
            ]
        )
        == 0
    )

    before = read_json(output / "exports" / "reqir.json")
    assert main(["review", str(output)]) == 0
    after = read_json(output / "exports" / "reqir.json")

    assert after["metadata"] == {
        **before["metadata"],
        "export_state": "review_snapshot",
    }


def test_review_cli_applies_actions_and_refreshes_outputs(tmp_path: Path):
    output = tmp_path / "demo"
    assert main(["extract", "docs/sample_srs.md", "--model-mode", "mock", "--output", str(output)]) == 0

    assert (
        main(
            [
                "review",
                str(output),
                "--id",
                "REQ-0001",
                "--action",
                "approve",
                "--reviewer",
                "local",
            ]
        )
        == 0
    )
    reqir = read_json(output / "exports" / "reqir.json")
    assert reqir["schema_version"] == "reqir_v4"
    evidence_fingerprint = read_json(output / "run_manifest.json")["evidence"][
        "evidence_fingerprint"
    ]
    assert reqir["metadata"] == {
        "export_state": "review_snapshot",
        "document": "sample_srs.md",
        "source_format": "markdown",
        "parser": "markdown_parser_v1",
        "evidence_fingerprint": evidence_fingerprint,
    }
    req_0001 = next(item for item in reqir["items"] if item["id"] == "REQ-0001")
    assert req_0001["review_status"] == "approved"

    tags_patch = tmp_path / "tags_patch.json"
    write_json(tags_patch, {"tags": ["user-management", "reviewed"]})
    assert (
        main(
            [
                "review",
                str(output),
                "--id",
                "REQ-0001",
                "--action",
                "edit",
                "--patch",
                str(tags_patch),
                "--reviewer",
                "local",
            ]
        )
        == 0
    )
    reqir = read_json(output / "exports" / "reqir.json")
    req_0001 = next(item for item in reqir["items"] if item["id"] == "REQ-0001")
    assert req_0001["review_status"] == "approved"
    assert req_0001["tags"] == ["user-management", "reviewed"]

    assert (
        main(
            [
                "review",
                str(output),
                "--id",
                "REQ-0002",
                "--action",
                "reject",
                "--reviewer",
                "local",
                "--reason",
                "not required",
            ]
        )
        == 0
    )
    reqir = read_json(output / "exports" / "reqir.json")
    req_0002 = next(item for item in reqir["items"] if item["id"] == "REQ-0002")
    assert req_0002["review_status"] == "rejected"

    statement_patch = tmp_path / "statement_patch.json"
    write_json(statement_patch, {"statement": "系统应记录完整的用户账号状态变更审计信息。"})
    assert (
        main(
            [
                "review",
                str(output),
                "--id",
                "REQ-0003",
                "--action",
                "edit",
                "--patch",
                str(statement_patch),
                "--reviewer",
                "local",
            ]
        )
        == 0
    )
    reqir = read_json(output / "exports" / "reqir.json")
    req_0003 = next(item for item in reqir["items"] if item["id"] == "REQ-0003")
    assert req_0003["review_status"] == "needs_recheck"

    assert (
        main(
            [
                "review",
                str(output),
                "--id",
                "REQ-0001",
                "--action",
                "reject",
                "--reviewer",
                "local",
            ]
        )
        == 0
    )
    assert (
        main(
            [
                "review",
                str(output),
                "--id",
                "REQ-0001",
                "--action",
                "restore",
                "--reviewer",
                "local",
            ]
        )
        == 0
    )
    reqir = read_json(output / "exports" / "reqir.json")
    req_0001 = next(item for item in reqir["items"] if item["id"] == "REQ-0001")
    assert req_0001["review_status"] == "pending"

    review_log = read_json(output / "review" / "review_log.json")
    assert len(review_log) >= 6
    assert {record["action"] for record in review_log} >= {"approve", "reject", "edit", "restore"}

    workbook = load_workbook(output / "exports" / "requirements.xlsx")
    assert workbook["Requirements"].max_row == len(reqir["items"]) + 1
