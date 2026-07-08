from pathlib import Path

from openpyxl import load_workbook

from spectrail.core.models import RequirementIR, SourceSpan
from spectrail.exporters.xlsx_exporter import HEADERS, export_requirements_xlsx


def test_xlsx_export_readable_with_fixed_headers(tmp_path: Path):
    requirement = RequirementIR(
        id="REQ-0001",
        title="审计",
        type="functional",
        ears_pattern="ubiquitous",
        statement="系统应记录事件。",
        sources=[
            SourceSpan(
                document_id="doc_001",
                section="审计",
                block_id="blk_0001",
                quote="系统应记录事件。",
                match_status="PASS_EXACT",
            )
        ],
        confidence=0.9,
        tags=["audit"],
    )
    path = tmp_path / "requirements.xlsx"
    export_requirements_xlsx([requirement], path)

    workbook = load_workbook(path)
    sheet = workbook["Requirements"]
    assert [cell.value for cell in sheet[1]] == HEADERS
    assert sheet.max_row == 2
    assert sheet["O2"].value
    assert sheet["P2"].value == "PASS_EXACT"
